"""
교보문고 새 엑셀 5개 vs 기존 DB 비교
- 새로 추가된 판매상품ID만 추출

사용법:
  python kyobo_diff.py
  → 입력: ../raw/ 엑셀 5개, ../db/교보문고_스크래핑결과.xlsx
  → 출력: 터미널에 추가된 ID 수 출력, ../temp/신규목록.xlsx
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

RAW_FILES = [
    ("../raw/확률과통계_교보문고_카테고리_상품리스트.xlsx",                      "확률과통계"),
    ("../raw/미적분2_교보문고_카테고리_상품리스트.xlsx",                         "미적분2"),
    ("../raw/공통수학1_공통수학2_대수_미적분1_교보문고_카테고리_상품리스트.xlsx", "공통수학1,공통수학2,대수,미적분1"),
    ("../raw/기하_교보문고_카테고리_상품리스트.xlsx",                            "기하"),
    ("../raw/수능_교보문고_카테고리_상품리스트.xlsx",                            "수능"),
]

DB_FILE      = "../db/교보문고_스크래핑결과.xlsx"
OUTPUT_FILE  = "../temp/신규목록.xlsx"

def save_excel(df, path):
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = 'A2'
    wb.save(path)

def main():
    result = {
        'raw_total': 0, 'db_total': 0, 'added': 0,
    }

    # 새 엑셀 5개 읽기
    new_dfs = []
    for filename, subject in RAW_FILES:
        df = pd.read_excel(filename)
        df['과목'] = subject
        new_dfs.append(df)
    new_df = pd.concat(new_dfs, ignore_index=True)
    new_ids = set(new_df['판매상품ID'].astype(str).str.strip())
    result['raw_total'] = len(new_df)
    print(f"새 엑셀 전체: {len(new_df)}개 ({len(new_ids)}개 고유 ID)")

    # 기존 DB 읽기
    try:
        db_df = pd.read_excel(DB_FILE)
        db_ids = set(db_df['판매상품ID'].astype(str).str.strip())
        result['db_total'] = len(db_df)
        print(f"기존 DB: {len(db_df)}개 ({len(db_ids)}개 고유 ID)")
    except FileNotFoundError:
        print("기존 DB 없음 → 전체 신규로 처리")
        db_ids = set()

    # 비교
    added_ids = new_ids - db_ids
    result['added'] = len(added_ids)
    print(f"\n추가된 ID: {len(added_ids)}개")

    if not added_ids:
        print("추가된 항목 없음")
        return result

    # 신규 항목 추출
    added_df = new_df[new_df['판매상품ID'].astype(str).str.strip().isin(added_ids)].reset_index(drop=True)
    save_excel(added_df, OUTPUT_FILE)
    print(f"→ {OUTPUT_FILE} ({len(added_df)}개)")
    return result

if __name__ == '__main__':
    main()