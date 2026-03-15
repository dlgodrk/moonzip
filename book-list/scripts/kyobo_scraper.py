"""
교보문고 문제집 디테일 스크래퍼 (전체 5개 파일)

사용법:
  pip install pandas openpyxl requests beautifulsoup4
  python kyobo_scraper.py
"""

import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DELAY = 0.3

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Accept-Language': 'ko-KR,ko;q=0.9',
    'Referer': 'https://product.kyobobook.co.kr/',
}

# ================================================================
# 파일명 → subject 매핑
# ================================================================
INPUT_FILES = [
    ("../raw/확률과통계_교보문고_카테고리_상품리스트.xlsx",                      "확률과통계"),
    ("../raw/미적분2_교보문고_카테고리_상품리스트.xlsx",                         "미적분2"),
    ("../raw/공통수학1_공통수학2_대수_미적분1_교보문고_카테고리_상품리스트.xlsx", "공통수학1,공통수학2,대수,미적분1"),
    ("../raw/기하_교보문고_카테고리_상품리스트.xlsx",                            "기하"),
    ("../raw/수능_교보문고_카테고리_상품리스트.xlsx",                            "수능"),
]

OUTPUT_FILE = "../db/교보문고_스크래핑결과.xlsx"

# ================================================================
# book_type / series_name 분류
# ================================================================
def nospace(s): return str(s).replace(" ", "").lower()

TYPE_유형서 = [
    "풍산자 필수유형","풍산자 유형기본서","풍산자 반복수학","풍산자 라이트유형",
    "짱 중요한 유형","짱 중요한 내신","짱 어려운 유형","짱 쉬운 유형",
    "일품","유형중심","유형반복R","유형만렙","유형 해결의 법칙",
    "유형 + 내신 고쟝이","유형 + 내신 고쟁이","완쏠 유형",
    "아름다운샘 Hi Math","쎈B","숨마쿰라우데 스타트업","수학입문",
    "수학의 바이블 유형ON","수력충전","베이직쎈",
    "만렙 PM","만렙 AM","라이트쎈","내신콘서트 고등수학 기출문제집",
    "개념원리 RPM","개념 + 유형","Hi Math",
    "EBS 올림포스 유형편","1등급 만들기",
    "짱 쉬운 내신","우공비Q+Q","Hexagon 헥사곤",
    "유형Zip","쎈","심플 자이스토리","신 수학의 바이블 BOB","수학중심","수매씽",
    "메가스터디 CPR","마플시너지","날선유형","기출의 파급효과",
    "기본개념과 실전연습 마더텅","교과서 다품","각(GAK)","EBS 올림포스","50일 수학",
]
TYPE_개념서 = [
    "한권으로 완성하는 수학","한권으로 시작하는 수학","풍산자","짤강","완쏠 개념",
    "실력 수학의 정석","숨마쿰라우데 수학 기본서","수학의 샘",
    "수학의 바이블 개념ON","메가헤르츠(Mhz)","메가헤르츠",
    "디딤돌수학 개념기본","기본 수학의 정석","개념원리","개념쎈 라이트","개념쎈",
    "개념루트","개념Zip","개념 해결의 법칙","개념 + 유형",
    "EBS 수학의 왕도",
    "연마수학","신 수학의 바이블","수학중심","수매씽 개념","메가헤르츠","매그넘","날선개념","마플교과서","50일 수학",
]
TYPE_기출 = [
    "한권으로 완성하는 기출","최강불패",
    "최강 3월 모의고사","착한기출",
    "짱 중요한 유형 수능 기출","짱 어려운 유형 수능 기출","짱 쉬운 유형 수능 기출",
    "짱 Final 실전모의고사","지피지기 백전백승","종로 핵파","씨뮬",
    "실전 마무리 봉투모의고사","수능한권","수능완승 EBS","수능기출 N회독",
    "수능 기출 올픽","너기출","내신콘서트 고등수학 기출문제집",
    "기출 Rebirth","개꿀수학",
    "N기출","Full수록","풀수록",
    "EBS 올림포스 전국연합학력평가 기출문제집","EBS 올림포스 고난도",
    "EBS 수능 기출의 미래","시대인재 수능기출","마더텅 수능기출문제집",
    "이동훈 기출문제집","100발 100중",
    "한석원의 기출분석","자이스토리 내신 핵심 기출","자이스토리","완자 기출픽",
    "수능 수학 해석법","수능 기출 각","마플수능기출총정리","기출의 파급효과","수능 하이엔드",
]
TYPE_EBS = [
    "실전 마무리 봉투모의고사","수능완승 EBS",
    "EBS 올림포스 유형편","EBS 올림포스 전국연합학력평가 기출문제집",
    "EBS 올림포스 고난도","EBS 수학의 왕도","EBS 수능특강","EBS 수능완성",
    "EBS NEXT Step","50일 수학",
]
TYPE_N제 = [
    "포카칩 N제","절대유형 N제","일격필살 N제",
    "어삼쉬사","메가스터디 N제","랑데뷰 N제","땅우 N-1제",
    "고난도 N제 수능대비 메디컬수학","지인선 N제","이해원 N제","규토 라이트 N제","이로운 N제",
    "한석원의 펀더멘털N제","한석원의 4의규칙","유형 정복하기 N제",
    "씨에스엠 107제","설맞이 아카이브","설레임 N제","샤인미 N제",
]
TYPE_심화서 = [
    "플래티넘","최강 TOT","절대등급","일품","일등급 수학","실력 수학의 정석",
    "시험직전R","수학의 신","블랙라벨","Hi High","1등급 만들기",
    "내신 하이엔드",
]

NORMALIZE_MAP = [
    ("유형.Zip",          "유형Zip"),
    ("개념.Zip",          "개념Zip"),
    ("Full수록",          "풀수록"),
    ("Xistory 자이스토리", "자이스토리"),
    ("메가헤르츠(Mhz)",   "메가헤르츠"),
    ("각(GAK)",           "각GAK"),
    ("각 GAK",            "각GAK"),
]

def normalize_title(title):
    for old, new in NORMALIZE_MAP:
        title = title.replace(old, new)
        title = title.replace(old.upper(), new)
        title = title.replace(old.lower(), new)
    return title

def has(kw_list, t):
    return any(nospace(kw) in t for kw in kw_list)

def classify_type(title):
    title = normalize_title(title)
    t = nospace(title)
    types = []
    if has(TYPE_유형서, t): types.append("유형서")
    if has(TYPE_개념서, t): types.append("개념서")
    if has(TYPE_심화서, t): types.append("심화서")
    if has(TYPE_기출, t):   types.append("기출문제집")
    if has(TYPE_N제, t):    types.append("N제")
    if has(TYPE_EBS, t):    types.append("EBS")
    if "유형서" in types and "개념서" in types: types.remove("개념서")
    if "개념서" in types and "심화서" in types: types.remove("심화서")
    return ", ".join(types) if types else ""

SERIES_LIST = []
for item in TYPE_유형서: SERIES_LIST.append((item, "유형서"))
for item in TYPE_개념서: SERIES_LIST.append((item, "개념서"))
for item in TYPE_기출:   SERIES_LIST.append((item, "기출문제집"))
for item in TYPE_EBS:    SERIES_LIST.append((item, "EBS"))
for item in TYPE_N제:    SERIES_LIST.append((item, "N제"))
for item in TYPE_심화서: SERIES_LIST.append((item, "심화서"))
SERIES_LIST = list(dict.fromkeys(SERIES_LIST))
SERIES_LIST.sort(key=lambda x: -len(x[0]))

def match_series(title):
    title = normalize_title(title)
    t = nospace(title)
    for name, _ in SERIES_LIST:
        if nospace(normalize_title(name)) in t:
            return name
    return ""


ORDERED_COLS = [
    '상품명','시리즈','유형','과목','저자','출시일','출판사','개정판여부','판매현황',
    '상품코드','판매상품ID','정가','판매가','할인율','적립율','적립포인트','링크','표지이미지','오류',
]

def save_excel(df, path):
    df = df.drop(columns=[c for c in ['순번'] if c in df.columns])
    ordered = [c for c in ORDERED_COLS if c in df.columns]
    rest = [c for c in df.columns if c not in ordered]
    df = df[ordered + rest]
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = 'A2'
    wb.save(path)
def scrape(product_id):
    url = f"https://product.kyobobook.co.kr/detail/{product_id}"
    result = {
        'url': url,
        'cover_image_url': None,
        'sale_status': None,
        'edition_status': '최신판',
        'error': None,
    }
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, 'html.parser')

        img_meta = soup.find('meta', attrs={'name': 'twitter:image'})
        if img_meta:
            result['cover_image_url'] = img_meta.get('content', '').strip()

        status_el = soup.select_one('div.prod_status_box span.status')
        result['sale_status'] = status_el.get_text(strip=True) if status_el else '구매가능'

        for div in soup.select('div.prod_info_text'):
            if '새로 출시된 개정판이 있습니다' in div.get_text():
                result['edition_status'] = '구판'
                break

    except requests.exceptions.HTTPError as e:
        result['error'] = f'HTTP {e.response.status_code}'
    except Exception as e:
        result['error'] = str(e)[:120]

    return result


# ================================================================
# 메인
# ================================================================
def main():
    all_dfs = []

    for filename, subject in INPUT_FILES:
        print(f"\n{'='*50}")
        print(f"파일: {filename}  (subject={subject})")
        print(f"{'='*50}")

        df = pd.read_excel(filename)
        total = len(df)

        df['과목']   = subject
        df['유형']   = df['상품명'].apply(classify_type)
        df['시리즈'] = df['상품명'].apply(match_series)

        results = []
        for i, row in df.iterrows():
            pid   = str(row['판매상품ID']).strip()
            title = str(row.get('상품명', ''))[:28]
            print(f"  [{i+1}/{total}] {title}... ", end='', flush=True)

            res = scrape(pid)
            results.append(res)

            if res['error']:
                print(f"ERROR: {res['error']}")
            else:
                print(f"{res['sale_status']} | {res['edition_status']}")

            time.sleep(DELAY)

        res_df = pd.DataFrame(results)
        df['링크']     = df['판매상품ID'].apply(lambda x: f"https://product.kyobobook.co.kr/detail/{str(x).strip()}")
        df['표지이미지'] = res_df['cover_image_url'].values
        df['판매현황']   = res_df['sale_status'].values
        df['개정판여부'] = res_df['edition_status'].values
        df['오류']       = res_df['error'].values

        df = df.rename(columns={
            '인물':        '저자',
            '발행(출시)일자': '출시일',
            '적립예정포인트': '적립포인트',
        })
        df = df.drop(columns=[c for c in ['분야', '구분'] if c in df.columns])

        all_dfs.append(df)

        interim = "../temp/" + filename.split("/")[-1].replace('.xlsx', '_스크래핑결과.xlsx')
        save_excel(df, interim)
        print(f"  → 중간저장: {interim}")

    final = pd.concat(all_dfs, ignore_index=True)
    save_excel(final, OUTPUT_FILE)
    print(f"\n전체 완료! → {OUTPUT_FILE}  (총 {len(final)}개)")


if __name__ == '__main__':
    main()