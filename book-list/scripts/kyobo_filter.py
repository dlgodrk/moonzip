"""
교보문고 스크래핑결과 필터링
- 절판 제외
- 구판 제외
- 시리즈 없는 것 제외

사용법:
  python kyobo_filter.py
  → 입력: 교보문고_스크래핑결과.xlsx
  → 출력: 교보문고_필터결과.xlsx
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

INPUT_FILE  = "../db/교보문고_스크래핑결과.xlsx"
OUTPUT_FILE = "../db/교보문고_필터결과.xlsx"

ORDERED_COLS = [
    '상품명','시리즈','유형','과목','저자','출시일','출판사','개정판여부','판매현황',
    '상품코드','판매상품ID','정가','판매가','할인율','적립율','적립포인트','링크','표지이미지','오류',
]

def save_excel(df, path):
    df = df.drop(columns=[c for c in ['순번'] if c in df.columns])
    ordered = [c for c in ORDERED_COLS if c in df.columns]
    rest = [c for c in df.columns if c not in ordered]
    df = df[ordered + rest]
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = 'A2'
    wb.save(path)

def main():
    df = pd.read_excel(INPUT_FILE)
    total = len(df)
    print(f"전체: {total}개")
    removed_list = []
    result = {'total': total, '절판': 0, '구판': 0, '시리즈없음': 0, '중복': 0, '최종': 0,
              '시리즈없음_목록': []}

    # 절판 제외
    cond_절판 = df['판매현황'].str.contains('절판', na=False)
    removed = df[cond_절판].copy(); removed['제거사유'] = '절판'
    removed_list.append(removed)
    result['절판'] = int(cond_절판.sum())
    df = df[~cond_절판].reset_index(drop=True)
    print(f"절판 제외: {cond_절판.sum()}개 → {len(df)}개 남음")

    # 구판 제외
    cond_구판 = df['개정판여부'] == '구판'
    removed = df[cond_구판].copy(); removed['제거사유'] = '구판'
    removed_list.append(removed)
    result['구판'] = int(cond_구판.sum())
    df = df[~cond_구판].reset_index(drop=True)
    print(f"구판 제외: {cond_구판.sum()}개 → {len(df)}개 남음")

    # 시리즈 없는 것 제외
    cond_no_series = df['시리즈'].isna() | (df['시리즈'] == '')
    removed = df[cond_no_series].copy(); removed['제거사유'] = '시리즈없음'
    removed_list.append(removed)
    result['시리즈없음'] = int(cond_no_series.sum())
    result['시리즈없음_목록'] = df[cond_no_series]['상품명'].tolist()
    df = df[~cond_no_series].reset_index(drop=True)
    print(f"시리즈 없음 제외: {cond_no_series.sum()}개 → {len(df)}개 남음")

    # 판매상품ID 중복 제거
    before_dedup = len(df)
    dup_ids = df[df.duplicated(subset='판매상품ID', keep='first')].copy()
    dup_ids['제거사유'] = 'ID중복'
    removed_list.append(dup_ids)
    df = df.drop_duplicates(subset='판매상품ID', keep='first').reset_index(drop=True)
    result['중복'] = before_dedup - len(df)
    result['최종'] = len(df)
    print(f"ID 중복 제거: {before_dedup - len(df)}개 → {len(df)}개 남음")

    save_excel(df, OUTPUT_FILE)
    print(f"\n완료! → {OUTPUT_FILE}  (총 {len(df)}개)")

    # 제거된 항목 저장
    removed_df = pd.concat(removed_list, ignore_index=True) if removed_list else pd.DataFrame()
    if not removed_df.empty:
        save_excel(removed_df, "../db/교보문고_제거목록.xlsx")
        print(f"제거목록 → ../db/교보문고_제거목록.xlsx  (총 {len(removed_df)}개)")

    # 시리즈 집계
    series_rows = []
    for (series, book_type), grp in df.groupby(['시리즈', '유형']):
        if not series:
            continue
        subjects = ','.join(sorted(set(
            s.strip() for cell in grp['과목'].dropna() for s in str(cell).split(',')
        )))
        publishers = grp['출판사'].value_counts()
        series_rows.append({
            '시리즈':  series,
            '유형':    book_type,
            '권수':    len(grp),
            '출판사':  publishers.index[0] if len(publishers) else '',
            '과목':    subjects,
        })
    series_df = pd.DataFrame(series_rows).sort_values(['유형', '시리즈']).reset_index(drop=True)
    series_df.to_excel('../db/시리즈목록.xlsx', index=False)
    from openpyxl import load_workbook as lw
    from openpyxl.utils import get_column_letter as gcl
    wb = lw('../db/시리즈목록.xlsx')
    ws = wb.active
    ws.auto_filter.ref = f"A1:{gcl(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = 'A2'
    wb.save('../db/시리즈목록.xlsx')
    print(f"시리즈목록 → ../db/시리즈목록.xlsx  ({len(series_df)}개 시리즈)")

    return result

if __name__ == '__main__':
    main()