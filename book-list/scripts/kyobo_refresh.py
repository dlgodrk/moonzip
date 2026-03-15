"""
교보문고 판매현황/개정판여부 갱신
- db/교보문고_스크래핑결과.xlsx 전체 재스크래핑 (판매현황, 개정판여부만)
- 주 1회 실행 권장

사용법:
  python kyobo_refresh.py
  → 입력: ../db/교보문고_스크래핑결과.xlsx
  → 출력: ../db/교보문고_스크래핑결과.xlsx (덮어쓰기)
"""

import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DB_FILE = "../db/교보문고_스크래핑결과.xlsx"
DELAY   = 0.3

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Accept-Language': 'ko-KR,ko;q=0.9',
    'Referer': 'https://product.kyobobook.co.kr/',
}

ORDERED_COLS = [
    '상품명','시리즈','유형','과목','저자','출시일','출판사','개정판여부','판매현황',
    '상품코드','판매상품ID','정가','판매가','할인율','적립율','적립포인트','링크','표지이미지','오류',
]

def save_excel(df, path):
    df = df.drop(columns=[c for c in ['순번'] if c in df.columns])
    ordered = [c for c in ORDERED_COLS if c in df.columns]
    rest = [c for c in df.columns if c not in ordered]
    df = df[ordered + rest]
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = 'A2'
    wb.save(path)

def scrape_status(product_id):
    url = f"https://product.kyobobook.co.kr/detail/{product_id}"
    result = {
        'sale_status':    None,
        'edition_status': '최신판',
        'error':          None,
    }
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, 'html.parser')

        status_el = soup.select_one('div.prod_status_box span.status')
        result['sale_status'] = status_el.get_text(strip=True) if status_el else '구매가능'

        for div in soup.select('div.prod_info_text'):
            if '새로 출시된 개정판이 있습니다' in div.get_text():
                result['edition_status'] = '구판'
                break

    except requests.exceptions.HTTPError as e:
        result['error'] = f'HTTP {e.response.status_code}'
    except Exception as e:
        result['error'] = str(e)[:120]

    return result

def main():
    df = pd.read_excel(DB_FILE)
    total = len(df)
    print(f"전체 {total}개 판매현황/개정판여부 갱신 시작\n")

    for i, row in df.iterrows():
        pid   = str(row['판매상품ID']).strip()
        title = str(row.get('상품명', ''))[:28]
        print(f"  [{i+1}/{total}] {title}... ", end='', flush=True)

        res = scrape_status(pid)

        df.at[i, '판매현황']   = res['sale_status']
        df.at[i, '개정판여부'] = res['edition_status']
        df.at[i, '오류']       = res['error']

        if res['error']:
            print(f"ERROR: {res['error']}")
        else:
            print(f"{res['sale_status']} | {res['edition_status']}")

        if (i + 1) % 100 == 0:
            save_excel(df, DB_FILE)
            print(f"  >>> 중간저장 완료 ({i+1}개)")

        time.sleep(DELAY)

    save_excel(df, DB_FILE)
    print(f"\n완료! → {DB_FILE}  (총 {total}개 갱신)")

if __name__ == '__main__':
    main()
